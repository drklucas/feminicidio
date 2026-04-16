"""
Gera os arquivos estáticos para o GitHub Pages.

Lê os .xlsx de archives/, compila os dados em memória e produz:
  docs/index.html            - HTML com dados estaduais embutidos como JS
  docs/data/{id}.json        - um JSON por município (carregado sob demanda)
  docs/.nojekyll             - desativa o Jekyll no GitHub Pages

Uso:
    pip install openpyxl
    python generate_static.py
"""

from __future__ import annotations
import csv
import json
import os
import re
import shutil
import unicodedata
from collections import defaultdict
from pathlib import Path

import openpyxl

# ── Caminhos ──────────────────────────────────────────────────────────────────
ROOT = Path(__file__).parent
ARCHIVES_DIR = ROOT / "archives"
TEMPLATE_HTML = ROOT / "artefacts" / "app" / "static" / "index.html"
DOCS_DIR = ROOT / "docs"
DATA_DIR = DOCS_DIR / "data"
CSV_GERAL_2012_2017 = (
    ARCHIVES_DIR
    / "12160259-site-violencia-contra-as-mulheres-2012-a-2017-atualizado-em-09-janeiro-2018-publicacao - Geral.csv"
)

# ── Mapeamento de abas ────────────────────────────────────────────────────────
ABA_MAP = {
    "feminicídio tentado":   "feminicidio_tentado",
    "femicídio tentado":     "feminicidio_tentado",
    "feminicídio consumado": "feminicidio_consumado",
    "femicídio consumado":   "feminicidio_consumado",
    "ameaça":                "ameaca",
    "estupro":               "estupro",
    "lesão corporal":        "lesao_corporal",
    "geral":                 "geral",
}

MESES = {
    "jan": 1, "fev": 2, "mar": 3, "abr": 4,
    "mai": 5, "jun": 6, "jul": 7, "ago": 8,
    "set": 9, "out": 10, "nov": 11, "dez": 12,
}
ARQUIVO_2012_2017_TOKEN = "2012-a-2017"
MESES_LONGOS = {
    "janeiro": 1,
    "fevereiro": 2,
    "marco": 3,
    "abril": 4,
    "maio": 5,
    "junho": 6,
    "julho": 7,
    "agosto": 8,
    "setembro": 9,
    "outubro": 10,
    "novembro": 11,
    "dezembro": 12,
}
TIPO_MAP_CSV = {
    "femicidio tentado": "feminicidio_tentado",
    "femicidio consumado": "feminicidio_consumado",
    "ameaca": "ameaca",
    "estupro": "estupro",
    "lesao corporal": "lesao_corporal",
    "geral": "geral",
}


# ── Leitura dos xlsx (mesma lógica de load_data.py, sem PostgreSQL) ───────────

def _extrair_ano(filename: str) -> int | None:
    m = re.search(r"(20\d{2})", filename)
    return int(m.group(1)) if m else None


def _arquivo_duplicado_nome(filename: str) -> bool:
    """Detecta cópias automáticas do Windows, ex.: 'arquivo (1).xlsx'."""
    return re.search(r"\(\d+\)\.xlsx$", filename.lower()) is not None


def _normalizar_txt(valor: str | None) -> str:
    if not valor:
        return ""
    base = unicodedata.normalize("NFKD", valor.strip().lower())
    return "".join(ch for ch in base if not unicodedata.combining(ch))


def _numero_ptbr(valor: str) -> int:
    v = valor.strip()
    if not v or v == "-":
        return 0
    v = v.replace(".", "").replace(",", ".")
    return int(float(v))


def _cab_mensal(headers) -> dict[int, int] | None:
    mapping = {}
    for i, h in enumerate(headers):
        if isinstance(h, str) and h.strip().lower() in MESES:
            mapping[i] = MESES[h.strip().lower()]
    return mapping or None


def _cab_anual(headers) -> dict[int, int] | None:
    mapping = {}
    for i, h in enumerate(headers):
        if isinstance(h, (int, float)) and 2000 <= int(h) <= 2030:
            mapping[i] = int(h)
    return mapping or None


def _ler_aba(ws, tipo_crime: str, ano_arquivo: int) -> list[tuple]:
    records = []
    col_mes = col_ano = municipio_col = None

    for row in ws.iter_rows(values_only=True):
        if col_mes is None and col_ano is None:
            for i, cell in enumerate(row):
                if isinstance(cell, str) and "munic" in cell.lower():
                    municipio_col = i
            if municipio_col is not None:
                col_mes = _cab_mensal(row)
                if col_mes is None:
                    col_ano = _cab_anual(row)
            continue

        if municipio_col is None:
            continue

        municipio = row[municipio_col] if municipio_col < len(row) else None
        if not municipio or not isinstance(municipio, str):
            continue
        municipio = municipio.strip().upper()
        if not municipio:
            continue
        _m = municipio.lower()
        if _m in ("município", "municipio", "total", "geral"):
            continue
        if len(municipio) > 100:
            continue
        if municipio[0].isdigit():
            continue

        if col_mes:
            for col_idx, mes in col_mes.items():
                if col_idx < len(row):
                    qtd = row[col_idx]
                    qtd = int(qtd) if isinstance(qtd, (int, float)) else 0
                    records.append((municipio, tipo_crime, ano_arquivo, mes, qtd))
        elif col_ano:
            for col_idx, ano in col_ano.items():
                if col_idx < len(row):
                    qtd = row[col_idx]
                    qtd = int(qtd) if isinstance(qtd, (int, float)) else 0
                    records.append((municipio, tipo_crime, ano, None, qtd))

    return records


def ler_todos_xlsx(archives_dir: Path) -> dict[tuple, int]:
    """
    Retorna um dict {(municipio, tipo_crime, ano, mes): quantidade}
    com deduplicação via primeira-escrita-ganha (igual ao ON CONFLICT DO NOTHING).
    """
    xlsx_files = sorted(
        f for f in os.listdir(archives_dir)
        if f.endswith(".xlsx") and not _arquivo_duplicado_nome(f)
    )
    print(f"Arquivos xlsx encontrados: {len(xlsx_files)}")

    unique: dict[tuple, int] = {}

    for filename in xlsx_files:
        filepath = archives_dir / filename
        ano_arquivo = _extrair_ano(filename)
        print(f"  -> {filename}  (ano base: {ano_arquivo})")

        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            tipo_crime = ABA_MAP.get(sheet_name.strip().lower())
            if tipo_crime is None:
                continue
            if tipo_crime == "geral" and ARQUIVO_2012_2017_TOKEN in filename.lower():
                # A aba Geral desse arquivo usa layout mensal do ano corrente e
                # não representa a série anual 2012-2017 por município.
                continue
            ws = wb[sheet_name]
            for rec in _ler_aba(ws, tipo_crime, ano_arquivo):
                key = rec[:4]  # (municipio, tipo, ano, mes)
                if key not in unique:
                    unique[key] = rec[4]
        wb.close()

    print(f"  Total de registros únicos: {len(unique):,}")
    return unique


# ── Agregações ────────────────────────────────────────────────────────────────

def agregar_estado(data: dict[tuple, int]):
    """Agrega no nível do estado (todos os municípios)."""
    # Anual: soma por (tipo, ano)  [inclui mes=None e mes=número]
    anual: dict[tuple, int] = defaultdict(int)
    # Mensal: soma por (ano, tipo, mes) onde mes IS NOT NULL
    mensal: dict[tuple, int] = defaultdict(int)

    for (mun, tipo, ano, mes), qtd in data.items():
        anual[(tipo, ano)] += qtd
        if mes is not None:
            mensal[(ano, tipo, mes)] += qtd

    anual_lista = [
        {"tipo": tipo, "ano": ano, "total": total}
        for (tipo, ano), total in sorted(anual.items(), key=lambda x: (x[0][1], x[0][0]))
    ]

    anos_com_mensal = sorted({ano for (ano, _, _) in mensal})
    mensal_por_ano: dict[str, list] = {}
    for ano in anos_com_mensal:
        mensal_por_ano[str(ano)] = [
            {"tipo": tipo, "mes": mes, "total": total}
            for (a, tipo, mes), total in sorted(mensal.items(), key=lambda x: (x[0][2], x[0][1]))
            if a == ano
        ]

    return anual_lista, mensal_por_ano


def agregar_por_municipio(data: dict[tuple, int]) -> dict[str, dict]:
    """Retorna {municipio: {anual: [...], mensal: {ano: [...]}}}."""
    mun_anual: dict[tuple, int] = defaultdict(int)
    mun_mensal: dict[tuple, int] = defaultdict(int)

    for (mun, tipo, ano, mes), qtd in data.items():
        mun_anual[(mun, tipo, ano)] += qtd
        if mes is not None:
            mun_mensal[(mun, ano, tipo, mes)] += qtd

    # Agrupa por município
    municipios: dict[str, dict] = {}

    for (mun, tipo, ano), total in mun_anual.items():
        m = municipios.setdefault(mun, {"anual": [], "mensal": {}})
        m["anual"].append({"tipo": tipo, "ano": ano, "total": total})

    for (mun, ano, tipo, mes), total in mun_mensal.items():
        m = municipios.setdefault(mun, {"anual": [], "mensal": {}})
        ano_str = str(ano)
        m["mensal"].setdefault(ano_str, []).append(
            {"tipo": tipo, "mes": mes, "total": total}
        )

    # Ordena listas internas
    for mun_data in municipios.values():
        mun_data["anual"].sort(key=lambda r: (r["ano"], r["tipo"]))
        for lst in mun_data["mensal"].values():
            lst.sort(key=lambda r: (r["mes"], r["tipo"]))

    return municipios


def agregar_ranking_municipios(data: dict[tuple, int], top_n: int = 20) -> dict[str, dict[str, list[dict]]]:
    """
    Retorna ranking anual por município para cada tipo:
    { "2024": { "estupro": [{"municipio": "...", "total": 123}, ...], ... }, ... }
    """
    ranking_base: dict[tuple, int] = defaultdict(int)
    for (mun, tipo, ano, _mes), qtd in data.items():
        ranking_base[(ano, tipo, mun)] += qtd

    agrupado: dict[tuple, list[tuple[str, int]]] = defaultdict(list)
    for (ano, tipo, mun), total in ranking_base.items():
        agrupado[(ano, tipo)].append((mun, total))

    ranking: dict[str, dict[str, list[dict]]] = {}
    for (ano, tipo), items in agrupado.items():
        top_items = sorted(items, key=lambda x: (-x[1], x[0]))[:top_n]
        ranking.setdefault(str(ano), {})[tipo] = [
            {"municipio": mun, "total": total}
            for mun, total in top_items
        ]

    return ranking


def carregar_mensal_geral_2012_2017(csv_path: Path) -> list[dict]:
    """
    Carrega mensal estadual do CSV agregado (2012-2017):
    [{"ano": ..., "mes": ..., "tipo": ..., "total": ...}, ...]
    """
    if not csv_path.exists():
        print(f"[warn] CSV agregado 2012-2017 nao encontrado: {csv_path.name}")
        return []

    registros: list[dict] = []
    ano_atual: int | None = None
    idx_mes: dict[int, int] = {}

    with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        for raw in reader:
            row = [c.strip() for c in raw]
            joined = " ".join(row)

            m = re.search(r"Ano de\s+(20\d{2})", joined)
            if m:
                ano_atual = int(m.group(1))
                idx_mes = {}
                continue

            if ano_atual is None:
                continue

            col_nome = row[1] if len(row) > 1 else ""
            nome_norm = _normalizar_txt(col_nome)

            if nome_norm == "municipio":
                for i, cell in enumerate(row):
                    mes = MESES_LONGOS.get(_normalizar_txt(cell))
                    if mes:
                        idx_mes[i] = mes
                continue

            tipo = TIPO_MAP_CSV.get(nome_norm)
            if not tipo or not idx_mes:
                continue

            for col_idx, mes in idx_mes.items():
                if col_idx >= len(row):
                    continue
                qtd = _numero_ptbr(row[col_idx])
                registros.append({
                    "ano": ano_atual,
                    "mes": mes,
                    "tipo": tipo,
                    "total": qtd,
                })

    return registros


def agregar_mensal_longo_rs(
    data: dict[tuple, int],
    csv_2012_2017: list[dict],
) -> list[dict]:
    """
    Série mensal contínua do RS:
    - 2012-2017: CSV agregado estadual
    - 2018+: soma de municípios vindos dos xlsx
    """
    mensal_2018: dict[tuple, int] = defaultdict(int)
    for (mun, tipo, ano, mes), qtd in data.items():
        if mes is None or ano < 2018:
            continue
        mensal_2018[(ano, mes, tipo)] += qtd

    registros = list(csv_2012_2017)
    for (ano, mes, tipo), total in sorted(mensal_2018.items(), key=lambda x: (x[0][0], x[0][1], x[0][2])):
        registros.append({
            "ano": ano,
            "mes": mes,
            "tipo": tipo,
            "total": total,
        })

    return registros


# ── Modificação do HTML ───────────────────────────────────────────────────────

def _injetar_dados(
    html: str,
    anual_lista: list,
    mensal_por_ano: dict,
    mensal_longo_rs: list,
    ranking_municipios: dict,
    anos: list,
    municipios_sorted: list,
    mun_idx: dict,
) -> str:
    """Substitui chamadas de API por dados embutidos e referências a arquivos estáticos."""

    # 1. Bloco de dados embutidos — inserido ANTES do <script> principal
    #    para que as variáveis estejam definidas quando init() for chamado.
    data_block = (
        "<script>\n"
        f"window.__ANUAL__     = {json.dumps(anual_lista, ensure_ascii=False)};\n"
        f"window.__ANOS__      = {json.dumps(anos, ensure_ascii=False)};\n"
        f"window.__MUNICIPIOS__= {json.dumps(municipios_sorted, ensure_ascii=False)};\n"
        f"window.__MENSAL__    = {json.dumps(mensal_por_ano, ensure_ascii=False)};\n"
        f"window.__MENSAL_LONGO_RS__ = {json.dumps(mensal_longo_rs, ensure_ascii=False)};\n"
        f"window.__RANKING__   = {json.dumps(ranking_municipios, ensure_ascii=False)};\n"
        f"window.__MUN_IDX__   = {json.dumps(mun_idx, ensure_ascii=False)};\n"
        "</script>\n"
    )
    # Ancora no primeiro caractere único do script principal
    SCRIPT_ANCHOR = "<script>\n// ── Config"
    html = html.replace(SCRIPT_ANCHOR, data_block + SCRIPT_ANCHOR, 1)

    # 2. init(): substitui Promise.all com os três fetch de estado
    html = html.replace(
        "const [ad, anos, municipios] = await Promise.all([\n"
        "    fetch('/api/anual').then(r => r.json()),\n"
        "    fetch('/api/anos').then(r => r.json()),\n"
        "    fetch('/api/municipios').then(r => r.json()),\n"
        "  ]);",
        "const [ad, anos, municipios] = await Promise.all([\n"
        "    Promise.resolve(window.__ANUAL__),\n"
        "    Promise.resolve(window.__ANOS__),\n"
        "    Promise.resolve(window.__MUNICIPIOS__),\n"
        "  ]);",
    )

    # 3. init(): fetch mensal do ano mais recente
    html = html.replace(
        "mensalData = await fetch(`/api/mensal?ano=${latestAno}`).then(r => r.json());",
        "mensalData = window.__MENSAL__[String(latestAno)] || [];",
    )

    # 4. selAno listener (gráfico mensal estado)
    html = html.replace(
        "    mensalData = await fetch(`/api/mensal?ano=${e.target.value}`).then(r => r.json());",
        "    mensalData = window.__MENSAL__[e.target.value] || [];",
    )

    # 5. selAnoTab listener (tabela mensal estado)
    html = html.replace(
        "    mensalData = await fetch(`/api/mensal?ano=${e.target.value}`).then(r => r.json());",
        "    mensalData = window.__MENSAL__[e.target.value] || [];",
    )

    # 6. buscarCidade(): substitui os dois fetch (anual e mensal por município)
    html = html.replace(
        "  anualCidadeData = await fetch(`/api/anual?municipio=${encodeURIComponent(nome)}`).then(r => r.json());",
        "  const __munIdx = window.__MUN_IDX__[nome];\n"
        "  const __munJson = __munIdx !== undefined\n"
        "    ? await fetch(`./data/${__munIdx}.json`).then(r => r.json())\n"
        "    : {anual: [], mensal: {}};\n"
        "  anualCidadeData = __munJson.anual;",
    )
    html = html.replace(
        "  mensalCidadeData = await fetch(`/api/mensal?ano=${latestAno}&municipio=${encodeURIComponent(nome)}`).then(r => r.json());",
        "  mensalCidadeData = __munJson.mensal[String(latestAno)] || [];",
    )

    # 7. selAnoCidade listener
    html = html.replace(
        "  mensalCidadeData = await fetch(\n"
        "    `/api/mensal?ano=${e.target.value}&municipio=${encodeURIComponent(cidadeSelecionada)}`\n"
        "  ).then(r => r.json());",
        "  const __idx = window.__MUN_IDX__[cidadeSelecionada];\n"
        "  if (__idx !== undefined) {\n"
        "    const __d = await fetch(`./data/${__idx}.json`).then(r => r.json());\n"
        "    mensalCidadeData = __d.mensal[e.target.value] || [];\n"
        "  } else {\n"
        "    mensalCidadeData = [];\n"
        "  }",
    )

    # 8. novo gráfico mensal contínuo do RS
    html = html.replace(
        "  mensalLongoData = await fetch('/api/geral-mensal-rs').then(r => r.json());",
        "  mensalLongoData = window.__MENSAL_LONGO_RS__ || [];",
    )

    return html


# ── Principal ─────────────────────────────────────────────────────────────────

def main():
    print("=== Gerando site estático para GitHub Pages ===\n")

    # Limpa e recria docs/
    if DOCS_DIR.exists():
        shutil.rmtree(DOCS_DIR)
    DATA_DIR.mkdir(parents=True)

    # 1. Lê os xlsx
    data = ler_todos_xlsx(ARCHIVES_DIR)

    # 2. Agrega
    print("\nAgregando dados do estado…")
    anual_lista, mensal_por_ano = agregar_estado(data)
    csv_mensal_2012_2017 = carregar_mensal_geral_2012_2017(CSV_GERAL_2012_2017)
    mensal_longo_rs = agregar_mensal_longo_rs(data, csv_mensal_2012_2017)
    ranking_municipios = agregar_ranking_municipios(data, top_n=20)

    print("Agregando dados por município…")
    municipios_data = agregar_por_municipio(data)

    municipios_sorted = sorted(municipios_data.keys())
    mun_idx = {m: i for i, m in enumerate(municipios_sorted)}
    anos = sorted({r["ano"] for r in anual_lista})

    print(f"  {len(municipios_sorted)} municípios  |  {len(anos)} anos  |  {len(anual_lista)} registros anuais")

    # 3. Escreve JSON por município
    print(f"\nEscrevendo {len(municipios_sorted)} arquivos de município em docs/data/…")
    for nome, idx in mun_idx.items():
        path = DATA_DIR / f"{idx}.json"
        path.write_text(
            json.dumps(municipios_data[nome], ensure_ascii=False, separators=(",", ":")),
            encoding="utf-8",
        )

    # 4. Gera index.html
    print("Gerando docs/index.html…")
    template = TEMPLATE_HTML.read_text(encoding="utf-8")
    html = _injetar_dados(
        template,
        anual_lista,
        mensal_por_ano,
        mensal_longo_rs,
        ranking_municipios,
        anos,
        municipios_sorted,
        mun_idx,
    )
    (DOCS_DIR / "index.html").write_text(html, encoding="utf-8")

    # 5. .nojekyll (desativa Jekyll no GitHub Pages)
    (DOCS_DIR / ".nojekyll").touch()

    print("\nConcluido. Arquivos gerados em docs/")
    print(f"  docs/index.html")
    print(f"  docs/data/0.json … docs/data/{len(municipios_sorted)-1}.json")


if __name__ == "__main__":
    main()
