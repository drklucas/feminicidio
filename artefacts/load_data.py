"""
ETL: lê os .xlsx de indicadores de violência contra mulheres (RS)
     e popula a tabela `ocorrencias` no PostgreSQL.

Pré-requisitos:
    pip install openpyxl psycopg2-binary

Uso:
    python load_data.py
"""

import os
import re
import openpyxl
import psycopg2
from psycopg2.extras import execute_values

# ── Configuração ──────────────────────────────────────────────────────────────
DB = dict(
    host="localhost",
    port=5432,
    dbname="feminicidio",
    user="postgres",
    password="postgres",
)

DATA_DIR = os.path.join(os.path.dirname(__file__), "..")  # pasta raiz com os .xlsx

# Mapeamento: nome da aba → chave normalizada
ABA_MAP = {
    "feminicídio tentado":  "feminicidio_tentado",
    "femicídio tentado":    "feminicidio_tentado",
    "feminicídio consumado": "feminicidio_consumado",
    "femicídio consumado":   "feminicidio_consumado",
    "ameaça":               "ameaca",
    "estupro":              "estupro",
    "lesão corporal":       "lesao_corporal",
    "geral":                "geral",
}

MESES = {"jan": 1, "fev": 2, "mar": 3, "abr": 4, "mai": 5, "jun": 6,
         "jul": 7, "ago": 8, "set": 9, "out": 10, "nov": 11, "dez": 12}

# ── Helpers ───────────────────────────────────────────────────────────────────

def extrair_ano_do_nome(filename: str) -> int | None:
    """Extrai o ano (20XX) do nome do arquivo."""
    m = re.search(r"(20\d{2})", filename)
    return int(m.group(1)) if m else None


def normalizar_aba(nome: str) -> str | None:
    return ABA_MAP.get(nome.strip().lower())


def cabecalho_mensal(headers) -> dict[int, int] | None:
    """
    Recebe a lista de valores do cabeçalho e retorna {col_index: numero_mes}
    se for um arquivo mensal, ou None se não reconhecer o formato.
    """
    mapping = {}
    for i, h in enumerate(headers):
        if isinstance(h, str) and h.strip().lower() in MESES:
            mapping[i] = MESES[h.strip().lower()]
    return mapping if mapping else None


def cabecalho_anual(headers) -> dict[int, int] | None:
    """
    Para o arquivo 2012-2017: retorna {col_index: ano}.
    """
    mapping = {}
    for i, h in enumerate(headers):
        if isinstance(h, (int, float)) and 2000 <= int(h) <= 2030:
            mapping[i] = int(h)
    return mapping if mapping else None


def ler_aba(ws, tipo_crime: str, ano_arquivo: int) -> list[tuple]:
    """
    Percorre a planilha e devolve lista de tuplas
    (municipio, tipo_crime, ano, mes, quantidade).
    Pula linhas sem município ou com cabeçalho.
    """
    records = []
    col_mes = None
    col_ano = None
    municipio_col = None

    for row in ws.iter_rows(values_only=True):
        # Procura a linha de cabeçalho (contém 'Município' ou 'Municipio')
        if col_mes is None and col_ano is None:
            for i, cell in enumerate(row):
                if isinstance(cell, str) and "munic" in cell.lower():
                    municipio_col = i
            if municipio_col is not None:
                col_mes = cabecalho_mensal(row)
                if col_mes is None:
                    col_ano = cabecalho_anual(row)
            continue  # pula a linha de cabeçalho ela mesma

        if municipio_col is None:
            continue  # ainda não achou cabeçalho

        municipio = row[municipio_col] if municipio_col < len(row) else None
        if not municipio or not isinstance(municipio, str):
            continue
        municipio = municipio.strip().upper()
        if not municipio:
            continue
        # filtra linhas de totais, cabeçalhos e notas de rodapé
        _m = municipio.lower()
        if _m in ("município", "municipio", "total", "geral"):
            continue
        if len(municipio) > 100:       # textos longos não são municípios
            continue
        if municipio[0].isdigit():     # notas de rodapé numeradas (ex: "1. ...")
            continue

        if col_mes:
            for col_idx, mes in col_mes.items():
                if col_idx < len(row):
                    qtd = row[col_idx]
                    qtd = int(qtd) if isinstance(qtd, (int, float)) else 0
                    records.append((municipio, tipo_crime, ano_arquivo, mes, qtd))

        elif col_ano:
            for col_idx, ano in col_ano.items():
                if col_idx < len(row):
                    qtd = row[col_idx]
                    qtd = int(qtd) if isinstance(qtd, (int, float)) else 0
                    records.append((municipio, tipo_crime, ano, None, qtd))

    return records


# ── ETL principal ─────────────────────────────────────────────────────────────

def main():
    xlsx_files = sorted(
        f for f in os.listdir(DATA_DIR) if f.endswith(".xlsx")
    )
    print(f"Arquivos encontrados: {len(xlsx_files)}")

    all_records: list[tuple] = []

    for filename in xlsx_files:
        filepath = os.path.join(DATA_DIR, filename)
        ano_arquivo = extrair_ano_do_nome(filename)
        print(f"\n→ {filename}  (ano base: {ano_arquivo})")

        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

        for sheet_name in wb.sheetnames:
            tipo_crime = normalizar_aba(sheet_name)
            if tipo_crime is None:
                print(f"   [skip] aba ignorada: {sheet_name!r}")
                continue

            ws = wb[sheet_name]
            records = ler_aba(ws, tipo_crime, ano_arquivo)
            print(f"   {sheet_name!r} → {tipo_crime}: {len(records)} registros")
            all_records.extend(records)

        wb.close()

    print(f"\nTotal de registros lidos: {len(all_records):,}")

    # ── Inserir no PostgreSQL ──────────────────────────────────────────────────
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    print("Inserindo no banco (ON CONFLICT DO NOTHING)…")

    BATCH = 5000
    inserted = 0
    for i in range(0, len(all_records), BATCH):
        batch = all_records[i : i + BATCH]
        execute_values(
            cur,
            """
            INSERT INTO ocorrencias (municipio, tipo_crime, ano, mes, quantidade)
            VALUES %s
            ON CONFLICT (municipio, tipo_crime, ano, mes) DO NOTHING
            """,
            batch,
        )
        inserted += cur.rowcount
        conn.commit()
        print(f"  {min(i + BATCH, len(all_records)):,} / {len(all_records):,} processados…")

    cur.close()
    conn.close()
    print(f"\nConcluído. {inserted:,} registros inseridos.")


if __name__ == "__main__":
    main()
